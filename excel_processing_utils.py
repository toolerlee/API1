import os
import csv
from openpyxl import Workbook
import gc

# Helper function to create Excel from CSV files
def _create_excel_from_csv_files(csv_input_directory, target_excel_path, headers_list, log_fn):
    log_fn(f"內部函數 (來自 excel_processing_utils)：開始從 {csv_input_directory} 的 CSV 檔案創建 Excel 到 {target_excel_path}")
    excel_workbook = Workbook()
    if 'Sheet' in excel_workbook.sheetnames: # Remove default sheet
        del excel_workbook['Sheet']
    
    csv_files_processed_count = 0
    has_created_any_sheet = False
    try:
        csv_filenames = [f for f in os.listdir(csv_input_directory) if f.endswith('.csv')]
        log_fn(f"  在 {csv_input_directory} 中找到 {len(csv_filenames)} 個 CSV 檔案準備處理。")

        if not csv_filenames:
            log_fn("  警告: 在指定目錄中未找到任何 CSV 檔案。")
            return None

        for csv_filename in csv_filenames:
            full_csv_path = os.path.join(csv_input_directory, csv_filename)
            sheet_name = os.path.splitext(csv_filename)[0][:31] 
            try:
                ws = excel_workbook.create_sheet(title=sheet_name)
                with open(full_csv_path, 'r', encoding='utf-8-sig') as csv_f:
                    csv_reader = csv.reader(csv_f)
                    for row_data in csv_reader: 
                        ws.append(row_data)
                log_fn(f"    已將 {csv_filename} 的內容寫入工作表 '{sheet_name}'")
                csv_files_processed_count += 1
                has_created_any_sheet = True
            except Exception as e_sheet_create:
                log_fn(f"    ❌ 處理 CSV 檔案 {csv_filename} 並創建工作表時發生錯誤: {e_sheet_create}")
        
        if has_created_any_sheet: 
            excel_workbook.save(target_excel_path)
            log_fn(f"✅ Excel 檔案已成功創建並儲存於 {target_excel_path} (處理了 {csv_files_processed_count} 個CSV檔案，創建了工作表)")
            return target_excel_path
        else:
            log_fn("警告: 沒有成功從任何 CSV 檔案創建工作表，Excel 未儲存。")
            return None 

    except Exception as e_main_csv_to_excel:
        log_fn(f"❌ 創建 Excel 檔案時發生主要錯誤: {str(e_main_csv_to_excel)}")
        # print(f"PYTHON_ERROR in _create_excel_from_csv_files: {e_main_csv_to_excel}") # Avoid direct print in util
        # import traceback # Avoid direct import in util
        # traceback.print_exc() # Avoid direct print in util
        # Consider re-raising or logging more details if this util is used in other contexts
        raise # Re-raise the exception to be caught by the caller in main_job
    finally:
        try:
            if 'excel_workbook' in locals(): del excel_workbook
            collected_csv_excel = gc.collect()
            log_fn(f"記憶體優化 (_create_excel_from_csv_files): gc.collect() 清理了 {collected_csv_excel} 個物件。")
        except Exception as e_gc_csv_excel:
            log_fn(f"記憶體優化 (_create_excel_from_csv_files) 時發生錯誤: {str(e_gc_csv_excel)}") 