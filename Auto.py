import tkinter as tk
from tkinter import messagebox
import os
import shutil
import glob
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from copy import copy
from collections import defaultdict
import re
from pathlib import Path
import threading
import requests
import zipfile
import io
import gdown
import json
from tkinter import filedialog

class RetryManagerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("自動化介面")
        self.root.geometry("600x700")  # 增加視窗大小
        
        # 設定主題和字體
        self.setup_theme()
        
        # 用於追蹤定時任務
        self.scheduled_task = None
        self.countdown_task = None
        self.countdown_seconds = 0
        self.is_scheduled = False
        
        # 創建主框架
        main_frame = tk.Frame(root, bg='#000000')  # 黑色背景
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 先建立按鈕
        self.create_buttons(main_frame)
        
        # 創建路徑顯示標籤
        self.path_label = tk.Label(
            main_frame, 
            text="", 
            wraplength=550, 
            justify=tk.LEFT,
            font=('Press Start 2P', 12),  # 增加字體大小到 12
            bg='#000000',
            fg='#FFFFFF'
        )
        self.path_label.pack(pady=8)
        
        # 創建帳號資訊標籤
        self.account_label = tk.Label(
            main_frame, 
            text="", 
            wraplength=550, 
            justify=tk.LEFT,
            font=('Press Start 2P', 12),  # 增加字體大小到 12
            bg='#000000',
            fg='#FFFFFF'
        )
        self.account_label.pack(pady=8)
        
        # 創建 retry 資訊標籤
        self.retry_label = tk.Label(
            main_frame, 
            text="", 
            wraplength=550, 
            justify=tk.LEFT,
            font=('Press Start 2P', 12),  # 增加字體大小到 12
            bg='#000000',
            fg='#FFFFFF'
        )
        self.retry_label.pack(pady=8)
        
        # 創建狀態標籤
        self.status_label = tk.Label(
            main_frame, 
            text="", 
            wraplength=550,
            font=('Press Start 2P', 12),  # 增加字體大小到 12
            bg='#000000',
            fg='#FFFFFF'
        )
        self.status_label.pack(pady=8)
        
        # 創建時鐘標籤（右下角）
        clock_frame = tk.Frame(root, bg='#000000')
        clock_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)
        self.clock_label = tk.Label(
            clock_frame, 
            text="", 
            font=('Press Start 2P', 12),  # 增加字體大小到 12
            bg='#000000',
            fg='#FFFFFF'
        )
        self.clock_label.pack(side=tk.RIGHT)
        
        # 用於追蹤 keyins 執行狀態
        self.keyins_running = False
        
        # 初始化顯示
        self.update_latest_path()
        self.update_account_info()
        self.update_retry_info()
        self.update_clock()  # 初始化時鐘
        
        # 設置定時更新
        self.root.after(1000, self.auto_update)
    
    def setup_theme(self):
        """設定主題和樣式"""
        # 設定按鈕樣式
        button_style = {
            'font': ('Press Start 2P', 12),  # 增加字體大小到 12
            'bg': '#4A4A4A',  # 深灰色背景
            'fg': '#FFFFFF',  # 白色文字
            'activebackground': '#666666',  # 按下時的背景色
            'activeforeground': '#FFFFFF',
            'relief': 'solid',  # 實線邊框
            'borderwidth': 2,
            'padx': 12,  # 增加水平內邊距
            'pady': 8    # 增加垂直內邊距
        }
        
        # 設定輸入框樣式
        entry_style = {
            'font': ('Press Start 2P', 8),
            'bg': '#000000',  # 黑色背景
            'fg': '#00FF00',  # 綠色文字（經典終端機風格）
            'relief': 'solid',
            'borderwidth': 2,
            'insertbackground': '#00FF00'  # 游標顏色
        }
        
        # 設定標籤樣式
        label_style = {
            'font': ('Press Start 2P', 8),
            'bg': '#000000',  # 黑色背景
            'fg': '#FFFFFF'   # 白色文字
        }
        
        # 設定視窗背景色
        self.root.configure(bg='#000000')
        
        # 保存樣式供後續使用
        self.button_style = button_style
        self.entry_style = entry_style
        self.label_style = label_style

    def create_buttons(self, parent):
        # 第一行按鈕框架
        button_frame1 = tk.Frame(parent, bg='#000000')
        button_frame1.pack(pady=5)
        
        # 執行 keyins 按鈕
        run_btn = tk.Button(
            button_frame1, 
            text="執行 keyins", 
            command=self.run_keyins,
            **self.button_style
        )
        run_btn.pack(side=tk.LEFT, padx=5)
        
        # 新增自動化按鈕
        auto_btn = tk.Button(
            button_frame1,
            text="一鍵運行",
            command=self.run_automation,
            **self.button_style
        )
        auto_btn.pack(side=tk.LEFT, padx=5)
        
        # 第二行按鈕框架
        button_frame2 = tk.Frame(parent, bg='#000000')
        button_frame2.pack(pady=5)
        
        # 打開最新 bonus 按鈕
        open_bonus_btn = tk.Button(
            button_frame2,
            text="打開最新 bonus",
            command=self.open_latest_bonus,
            **self.button_style
        )
        open_bonus_btn.pack(side=tk.LEFT, padx=5)

        # 打開 config 按鈕
        open_config_btn = tk.Button(
            button_frame2,
            text="打開 config",
            command=self.open_config,
            **self.button_style
        )
        open_config_btn.pack(side=tk.LEFT, padx=5)

        # 打開 account 按鈕
        open_account_btn = tk.Button(
            button_frame2,
            text="打開 account",
            command=self.open_account,
            **self.button_style
        )
        open_account_btn.pack(side=tk.LEFT, padx=5)

        # 打開最新日誌按鈕
        open_log_btn = tk.Button(
            button_frame2,
            text="打開最新日誌",
            command=self.open_latest_log,
            **self.button_style
        )
        open_log_btn.pack(side=tk.LEFT, padx=5)

        # 第四行按鈕框架
        button_frame4 = tk.Frame(parent, bg='#000000')
        button_frame4.pack(pady=5)

        # 驗算工作表數量按鈕
        count_sheets_btn = tk.Button(
            button_frame4,
            text="驗算工作表數量",
            command=self.count_sheets,
            **self.button_style
        )
        count_sheets_btn.pack(side=tk.LEFT, padx=5)

        # 生成報表按鈕
        generate_report_btn = tk.Button(
            button_frame4,
            text="生成報表",
            command=self.generate_report,
            **self.button_style
        )
        generate_report_btn.pack(side=tk.LEFT, padx=5)

        # 第五行按鈕框架（定時設定）
        button_frame5 = tk.Frame(parent, bg='#000000')
        button_frame5.pack(pady=5)

        # 時間設定標籤
        time_label = tk.Label(
            button_frame5, 
            text="自動化定時：",
            font=('Press Start 2P', 12),  # 增加字體大小
            bg='#000000',
            fg='#FFFFFF'
        )
        time_label.pack(side=tk.LEFT, padx=5)

        # 讀取保存的時間設定
        saved_hour, saved_minute = self.load_schedule_time()

        # 小時輸入框
        hour_frame = tk.Frame(button_frame5, bg='#000000')
        hour_frame.pack(side=tk.LEFT, padx=2)
        hour_label = tk.Label(
            hour_frame, 
            text="[",
            font=('Press Start 2P', 12),  # 增加字體大小
            bg='#000000',
            fg='#FFFFFF'
        )
        hour_label.pack(side=tk.LEFT)
        
        self.hour_var = tk.StringVar(value=saved_hour.zfill(2))
        hour_entry = tk.Entry(
            hour_frame, 
            textvariable=self.hour_var, 
            width=2, 
            justify='center',
            font=('Press Start 2P', 12),  # 增加字體大小
            bg='#000000',
            fg='#00FF00',
            relief='solid',
            borderwidth=2,
            insertbackground='#00FF00'
        )
        hour_entry.pack(side=tk.LEFT)
        
        hour_label2 = tk.Label(
            hour_frame, 
            text="]",
            font=('Press Start 2P', 12),  # 增加字體大小
            bg='#000000',
            fg='#FFFFFF'
        )
        hour_label2.pack(side=tk.LEFT)

        # 冒號標籤
        colon_label = tk.Label(
            button_frame5, 
            text=":",
            font=('Press Start 2P', 12),  # 增加字體大小
            bg='#000000',
            fg='#FFFFFF'
        )
        colon_label.pack(side=tk.LEFT)

        # 分鐘輸入框
        minute_frame = tk.Frame(button_frame5, bg='#000000')
        minute_frame.pack(side=tk.LEFT, padx=2)
        minute_label = tk.Label(
            minute_frame, 
            text="[",
            font=('Press Start 2P', 12),  # 增加字體大小
            bg='#000000',
            fg='#FFFFFF'
        )
        minute_label.pack(side=tk.LEFT)
        
        self.minute_var = tk.StringVar(value=saved_minute.zfill(2))
        minute_entry = tk.Entry(
            minute_frame, 
            textvariable=self.minute_var, 
            width=2, 
            justify='center',
            font=('Press Start 2P', 12),  # 增加字體大小
            bg='#000000',
            fg='#00FF00',
            relief='solid',
            borderwidth=2,
            insertbackground='#00FF00'
        )
        minute_entry.pack(side=tk.LEFT)
        
        minute_label2 = tk.Label(
            minute_frame, 
            text="]",
            font=('Press Start 2P', 12),  # 增加字體大小
            bg='#000000',
            fg='#FFFFFF'
        )
        minute_label2.pack(side=tk.LEFT)

        # 定時啟動按鈕
        self.schedule_btn = tk.Button(
            button_frame5,
            text="定時啟動",
            command=self.toggle_schedule,
            width=15,
            height=2,
            font=('Press Start 2P', 12),  # 增加字體大小
            bg='#4A4A4A',
            fg='#FFFFFF',
            activebackground='#666666',
            activeforeground='#FFFFFF',
            relief='solid',
            borderwidth=2,
            padx=12,
            pady=8
        )
        self.schedule_btn.pack(side=tk.LEFT, padx=10)

        # 綁定驗證函數
        def validate_time(P):
            if P == "": return True
            if len(P) > 2: return False
            if not P.isdigit(): return False
            if len(P) == 2:
                if P[0] == "0" and P[1] == "0": return True
                if P[0] == "0" and P[1] != "0": return True
                if P[0] == "1" and P[1] in "0123456789": return True
                if P[0] == "2" and P[1] in "0123": return True
                return False
            return True

        def validate_minute(P):
            if P == "": return True
            if len(P) > 2: return False
            if not P.isdigit(): return False
            if len(P) == 2:
                if P[0] == "0" and P[1] == "0": return True
                if P[0] == "0" and P[1] != "0": return True
                if P[0] == "1" and P[1] in "0123456789": return True
                if P[0] == "2" and P[1] in "0123456789": return True
                if P[0] == "3" and P[1] in "0123456789": return True
                if P[0] == "4" and P[1] in "0123456789": return True
                if P[0] == "5" and P[1] in "0123456789": return True
                return False
            return True

        vcmd = (self.root.register(validate_time), '%P')
        vcmd_minute = (self.root.register(validate_minute), '%P')
        hour_entry.config(validate='key', validatecommand=vcmd)
        minute_entry.config(validate='key', validatecommand=vcmd_minute)

        # 新增第七行按鈕框架（更新按鈕）
        button_frame7 = tk.Frame(parent, bg='#F0F0F0')
        button_frame7.pack(pady=5)

        # 更新按鈕樣式修改
        self.update_btn = tk.Button(
            button_frame7,
            text="更新/下載",
            command=self.check_update,
            state=tk.DISABLED,
            width=10,
            height=1,
            **self.button_style
        )
        self.update_btn.pack(side=tk.LEFT, padx=5)

    def update_clock(self):
        """更新時鐘顯示"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.clock_label.config(text=current_time)
        self.root.after(1000, self.update_clock)  # 每秒更新一次

    def auto_update(self):
        """定時更新所有資訊"""
        self.update_latest_path()
        self.update_account_info()
        self.update_retry_info()
        self.root.after(1000, self.auto_update)  # 設置下一次更新

    def update_retry_info(self):
        """更新 retry 資訊顯示"""
        try:
            # 獲取最新的 retry.txt
            log_dirs = glob.glob("logs/*")
            if not log_dirs:
                self.retry_label.config(text="找不到 logs 目錄")
                return
            
            latest_log_dir = max(log_dirs, key=os.path.getctime)
            retry_file = os.path.join(latest_log_dir, "retry.txt")
            
            if not os.path.exists(retry_file):
                self.retry_label.config(text="目前無 retry.txt 文件")
                return
            
            # 獲取文件修改時間
            mod_time = datetime.fromtimestamp(os.path.getmtime(retry_file))
            mod_time_str = mod_time.strftime("%Y-%m-%d %H:%M:%S")
            
            # 讀取 retry.txt 的內容
            with open(retry_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            # 計算帳號數量（每三行為一個帳號）
            account_count = len(lines) // 3
            
            # 顯示信息
            info = f"最新 retry 文件信息：\n"
            info += f"路徑：{retry_file}\n"
            info += f"修改時間：{mod_time_str}\n"
            info += f"失敗帳號數量：{account_count} 個"
            
            self.retry_label.config(text=info)
            
        except Exception as e:
            self.retry_label.config(text=f"錯誤: {str(e)}")

    def update_account_info(self):
        """更新帳號資訊顯示"""
        try:
            if not os.path.exists("account.txt"):
                self.account_label.config(text="找不到 account.txt 文件")
                return
            
            # 讀取 account.txt 的內容
            with open("account.txt", 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            # 計算帳號數量（每三行為一個帳號）
            account_count = len(lines) // 3
            
            # 顯示簡潔的結果
            self.account_label.config(text=f"account.txt 帳號數量：{account_count}個")
            
        except Exception as e:
            self.account_label.config(text=f"錯誤: {str(e)}")

    def run_keyins(self):
        if self.keyins_running:
            self.status_label.config(text="keyins 正在執行中，請稍候...")
            return
            
        def run_keyins_thread():
            try:
                self.keyins_running = True
                
                # 檢查並生成所有必要檔案
                missing_files = []
                
                # 1. 檢查 config.txt
                if not os.path.exists("config.txt"):
                    missing_files.append("config.txt")
                    config_content = (
                        "# 模式設定\n"
                        "mode=0  # 0=一般模式, 1=重試模式\n\n"
                        "# 並發處理設定\n"
                        "max_concurrent_accounts=10  # 同時處理的帳號數量\n"
                        "thread_start_delay=0.5  # 線程啟動延遲(秒)\n\n"
                        "# 登入重試設定\n"
                        "max_login_attempts=3  # 登入重試次數\n\n"
                        "# 請求間隔設定\n"
                        "request_delay=2.0  # 每次請求之間的延遲時間(秒)\n"
                        "max_request_retries=3  # 請求失敗時的最大重試次數\n"
                        "retry_delay=5.0  # 重試之間的延遲時間(秒)\n\n"
                        "# 日期設定\n"
                        "start_date=2025/01/01  # 開始日期\n"
                        "end_date=2025/12/31  # 結束日期\n"
                    )
                    with open("config.txt", 'w', encoding='utf-8') as f:
                        f.write(config_content)
                # 2. 檢查 account.txt
                if not os.path.exists("account.txt"):
                    missing_files.append("account.txt")
                    with open("account.txt", 'w', encoding='utf-8') as f:
                        pass
                # 3. 檢查 terminal_messages.txt
                if not os.path.exists("terminal_messages.txt"):
                    missing_files.append("terminal_messages.txt")
                    terminal_messages_content = """[SHOW]
# 帳號處理狀態
[線程 - {name} ({account})] 所有數據收集完成，已加入寫入隊列

# 錯誤與警告
[線程 - {name} ({account})] 登入失敗原因: 帳號密碼錯誤或已被停權
[線程 - {name} ({account})] 此錯誤可能是由於系統效能吃緊導致，若次數頻繁請降低併發帳號數量或檢查效能資源。
[線程 - {name} ({account})] 所有連接和登入嘗試均失敗。該帳號處理終止。
未找到獎金歷史總計行，已添加佔位符

# 進度報告
處理結果詳細總結
總帳號數 (本輪)：
成功處理 (加入Excel隊列)：
登入失敗：
資料擷取失敗：
Excel寫入/準備失敗：
線程錯誤 (嚴重)：
偵測到新帳號(無歷史資料)：
本輪處理耗時：
本輪 ({total_count} 個帳號) 處理完成。

[HIDE]
# 調試訊息
Browser created for attempt
初始化成功
辨識驗證碼
ChromeDriver service
準備啟動處理帳號
嘗試第
DEBUG_ROWS
DEBUG_KEYBOARD

# 一般進度訊息
未偵測到 '無資料' 彈窗，繼續檢查獎金歷史表格...
點擊查詢按鈕...
等待瀏覽器初始化完成
瀏覽器初始化完成
已點擊錯誤彈窗上的確認按鈕
錯誤彈窗已成功關閉
已嘗試使用 ESC 鍵關閉未消失的彈窗
ESC 鍵成功使彈窗關閉
已為 '無資料(彈窗)' 的情況準備預設獎金歷史數據列
已成功合併個人和會員資料到預設/空獎金歷史
獎金歷史表格，找到
ChromeDriver service 已停止。
[線程 - {name} ({account})] 登入成功！
[線程 - {name} ({account})] 瀏覽器已關閉。
等待瀏覽器初始化完成
秒後啟動線程...
偵測到 '總計' 行，將單獨處理。
158|瀏覽器已關閉。"""
                    with open("terminal_messages.txt", 'w', encoding='utf-8') as f:
                        f.write(terminal_messages_content)
                # 如果有任何檔案被生成，顯示提示訊息
                if missing_files:
                    files_str = "、".join(missing_files)
                    self.status_label.config(text=f"已生成以下檔案：{files_str}\n請檢查設定後重新執行")
                    return
                # 讀取並顯示帳號數量
                with open("account.txt", 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                account_count = len(lines) // 3
                self.status_label.config(text=f"共:{account_count}個帳號\nkeyins 執行中...")
                # 優先以 exe 執行，若不存在則執行 py
                if os.path.exists("keyins.exe"):
                    os.system("keyins.exe")
                else:
                    os.system("python keyins.py")
                self.status_label.config(text=f"共:{account_count}個帳號\nkeyins 執行完成")
            except Exception as e:
                self.status_label.config(text=f"錯誤: {str(e)}")
            finally:
                self.keyins_running = False
        # 在背景線程中執行 keyins
        thread = threading.Thread(target=run_keyins_thread)
        thread.daemon = True
        thread.start()

    def generate_report(self):
        def report_thread():
            try:
                import openpyxl
                import os
                import re
                from collections import defaultdict
                from openpyxl.styles import Alignment, Font, Border, Side
                from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
                from copy import copy

                # === 自動搜尋 bonus.xlsx ===
                import glob
                import os
                bonus_files = glob.glob("output/**/bonus.xlsx", recursive=True)
                bonus_files += glob.glob("output/**/Bonus.xlsx", recursive=True)
                bonus_files += glob.glob("bonus.xlsx")
                bonus_files += glob.glob("Bonus.xlsx")
                if not bonus_files:
                    self.status_label.config(text="找不到 bonus.xlsx 或 Bonus.xlsx 檔案")
                    return
                bonus_path = max(bonus_files, key=os.path.getmtime)
                bonus_dir = os.path.dirname(bonus_path)
                bonus2_path = os.path.join(bonus_dir, "Bonus2.xlsx")
                if os.path.exists(bonus2_path):
                    os.remove(bonus2_path)
                wb_source = openpyxl.load_workbook(bonus_path, data_only=True)
                wb_target = openpyxl.Workbook()
                wb_target.remove(wb_target.active)

                person_sheets = defaultdict(list)
                for sheet_name_from_bonus_xlsx in wb_source.sheetnames:
                    name_raw_part = sheet_name_from_bonus_xlsx.split("_")[0]
                    person_identifier_for_grouping = re.sub(r'\d+', '', name_raw_part)
                    person_sheets[person_identifier_for_grouping].append(sheet_name_from_bonus_xlsx)

                color_map = {
                    "紅利積分": "FF0000",
                    "電子錢包": "00008B",
                    "獎金暫存": "8B4513",
                    "註冊分": "FF8C00",
                    "商品券": "2F4F4F",
                    "星級": "708090"
                }
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                def is_number(value):
                    if value is None:
                        return False
                    try:
                        float(str(value).replace(',', ''))
                        return True
                    except (ValueError, TypeError):
                        return False
                def sort_sheets_by_gold_level(sheet_names, wb_source):
                    def get_sheet_order(sheet_name):
                        ws = wb_source[sheet_name]
                        star_level = ws['V2'].value
                        return 1 if star_level and "金級" in str(star_level) else 0
                    return sorted(sheet_names, key=get_sheet_order)
                def apply_border_to_cell(cell):
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    original_font = cell.font or Font()
                    cell.font = Font(
                        name=original_font.name,
                        size=original_font.size,
                        bold=True,
                        italic=original_font.italic,
                        vertAlign=original_font.vertAlign,
                        underline=original_font.underline,
                        strike=original_font.strike,
                        color=original_font.color
                    )
                # === 日期對齊修正（只用 all_dates，且只寫一次） ===
                # 1. 收集所有日期
                all_dates_set = set()
                for person_id, source_sheet_names_list in person_sheets.items():
                    for s_name in source_sheet_names_list:
                        ws_s = wb_source[s_name]
                        dates = [row[0] for row in ws_s.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True) if row[0] is not None]
                        all_dates_set.update(dates)
                all_dates = sorted(all_dates_set, reverse=True)  # 最新日期在上
                # 2. 主循環：為每個 person_identifier 創建一個工作表
                for person_id, source_sheet_names_list in person_sheets.items():
                    sorted_sheet_names = sort_sheets_by_gold_level(source_sheet_names_list, wb_source)
                    ws_target = wb_target.create_sheet(title=person_id)
                    ws_target.insert_rows(2, amount=5)
                    ws_target['A1'] = person_id
                    ws_target['A1'].font = Font(bold=True)
                    apply_border_to_cell(ws_target['A1'])
                    if not sorted_sheet_names:
                        continue
                    first_source_sheet_name = sorted_sheet_names[0]
                    ws_first_source = wb_source[first_source_sheet_name]
                    q1_value = ws_first_source['Q1'].value
                    ws_target['A2'] = q1_value
                    font_color_a2 = color_map.get(str(q1_value).strip(), None)
                    if font_color_a2:
                        ws_target['A2'].font = Font(color=font_color_a2)
                    apply_border_to_cell(ws_target['A2'])
                    for idx, col_letter in enumerate(['R', 'S', 'T', 'U', 'V']):
                        value = ws_first_source[f'{col_letter}1'].value
                        cell = ws_target[f'A{3 + idx}']
                        cell.value = value
                        font_color = color_map.get(str(value).strip(), None)
                        if font_color:
                            cell.font = Font(color=font_color)
                        apply_border_to_cell(cell)
                    # 寫入主日期欄（只寫一次）
                    for idx, date_val in enumerate(all_dates):
                        cell = ws_target.cell(row=11 + idx, column=1, value=date_val)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        apply_border_to_cell(cell)
                    # 補齊A2~A7的數據（Q2~V2）
                    for sheet_idx, current_source_sheet_name in enumerate(sorted_sheet_names):
                        ws_current_source = wb_source[current_source_sheet_name]
                        target_col_for_this_source = 2 + sheet_idx
                        # 先補A2~A7（Q2~V2）
                        for row_idx, col_letter in enumerate(['Q', 'R', 'S', 'T', 'U', 'V']):
                            value = ws_current_source[f'{col_letter}2'].value
                            cell = ws_target.cell(row=2 + row_idx, column=target_col_for_this_source)
                            if is_number(value):
                                cell.value = float(str(value).replace(',', ''))
                                cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                            else:
                                cell.value = value
                            # 設定顏色
                            row_title = ws_target[f'A{2 + row_idx}'].value
                            font_color = color_map.get(str(row_title).strip(), None)
                            if font_color:
                                cell.font = Font(color=font_color)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            apply_border_to_cell(cell)
                        # 其餘原本M欄日期對齊數據填入（根據 all_dates 對齊，只寫一次）
                        date_to_m = {}
                        for row in ws_current_source.iter_rows(min_row=2, max_col=13, values_only=True):
                            date = row[0]
                            m_val = row[12]
                            if date is not None:
                                date_to_m[date] = m_val
                        for idx, date_val in enumerate(all_dates):
                            m_val = date_to_m.get(date_val, None)
                            cell = ws_target.cell(row=11 + idx, column=target_col_for_this_source)
                            if is_number(m_val):
                                cell.value = float(str(m_val).replace(',', ''))
                                cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                            else:
                                cell.value = m_val
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            apply_border_to_cell(cell)
                    # 其餘內容保持不變，移除 max_dates 相關貼資料的部分
                    # A10（總計）B~帳號欄標註紫色
                    num_data_columns = len(sorted_sheet_names)
                    for col in range(2, 2 + num_data_columns):
                        cell = ws_target.cell(row=10, column=col)
                        original_font = cell.font or Font()
                        cell.font = Font(
                            name=original_font.name,
                            size=original_font.size,
                            bold=True,
                            italic=original_font.italic,
                            vertAlign=original_font.vertAlign,
                            underline=original_font.underline,
                            strike=original_font.strike,
                            color="8B008B"
                        )
                        apply_border_to_cell(cell)
                    max_dates = []
                    for s_name in sorted_sheet_names:
                        ws_s = wb_source[s_name]
                        dates_from_source_a_col = [row[0] for row in ws_s.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True) if row[0] is not None]
                        if len(dates_from_source_a_col) > len(max_dates):
                            max_dates = dates_from_source_a_col
                    for idx, date_val in enumerate(max_dates):
                        cell = ws_target.cell(row=11 + idx, column=1, value=date_val)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        apply_border_to_cell(cell)
                    for sheet_idx, current_source_sheet_name in enumerate(sorted_sheet_names):
                        ws_current_source = wb_source[current_source_sheet_name]
                        target_col_for_this_source = 2 + sheet_idx
                        source_m_column_data = [row[0] for row in ws_current_source.iter_rows(min_row=2, min_col=13, max_col=13, values_only=True) if row[0] is not None]
                        if source_m_column_data:
                            total_value_from_m = source_m_column_data[-1]
                            total_cell_target = ws_target.cell(row=10, column=target_col_for_this_source)
                            if is_number(total_value_from_m):
                                total_cell_target.value = float(str(total_value_from_m).replace(',', ''))
                                total_cell_target.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                            else:
                                total_cell_target.value = total_value_from_m
                            row_title = ws_target[f'A{2 + (target_col_for_this_source-2)}'].value if 2 <= (2 + (target_col_for_this_source-2)) <= 7 else None
                            font_color = color_map.get(str(row_title).strip(), None) if row_title else None
                            if font_color:
                                total_cell_target.font = Font(color=font_color, bold=True)
                            apply_border_to_cell(total_cell_target)
                            for m_data_idx, m_val in enumerate(source_m_column_data[:-1]):
                                if 11 + m_data_idx <= 10 + len(max_dates):
                                    cell_m_data_target = ws_target.cell(row=11 + m_data_idx, column=target_col_for_this_source)
                                    if is_number(m_val):
                                        cell_m_data_target.value = float(str(m_val).replace(',', ''))
                                        cell_m_data_target.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                                    else:
                                        cell_m_data_target.value = m_val
                                    cell_m_data_target.alignment = Alignment(horizontal='center', vertical='center')
                                    apply_border_to_cell(cell_m_data_target)
                        else:
                            apply_border_to_cell(ws_target.cell(row=10, column=target_col_for_this_source))
                            for m_data_idx_empty in range(len(max_dates)):
                                if 11 + m_data_idx_empty <= 10 + len(max_dates):
                                    apply_border_to_cell(ws_target.cell(row=11 + m_data_idx_empty, column=target_col_for_this_source))
                                    ws_target.cell(row=11 + m_data_idx_empty, column=target_col_for_this_source).alignment = Alignment(horizontal='center', vertical='center')
                        w_val = ws_current_source['W2'].value
                        x_val = ws_current_source['X2'].value
                        lr_cell_target = ws_target.cell(row=9, column=target_col_for_this_source)
                        if w_val is not None and x_val is not None:
                            lr_cell_target.value = f"{w_val} <> {x_val}"
                        lr_cell_target.alignment = Alignment(horizontal='center', vertical='center')
                        lr_cell_target.font = Font(color="006400")
                        apply_border_to_cell(lr_cell_target)
                        if sheet_idx == 0:
                            ws_target["A8"] = "帳號"
                            apply_border_to_cell(ws_target["A8"])
                        account_name = current_source_sheet_name.split('_', 1)[-1] if '_' in current_source_sheet_name else current_source_sheet_name
                        name_part = current_source_sheet_name.split('_')[0]
                        account_cell_target = ws_target.cell(row=8, column=target_col_for_this_source, value=account_name)
                        account_cell_target.font = Font(color="000000")
                        apply_border_to_cell(account_cell_target)
                        if sheet_idx == 0:
                            ws_target["A1"] = "名稱"
                            apply_border_to_cell(ws_target["A1"])
                        name_cell_target = ws_target.cell(row=1, column=target_col_for_this_source, value=name_part)
                        name_cell_target.font = Font(bold=True)
                        apply_border_to_cell(name_cell_target)
                    ws_target.column_dimensions['A'].width = 11
                    num_data_columns = len(sorted_sheet_names)
                    for col_num_idx in range(num_data_columns):
                        col_letter_dynamic = openpyxl.utils.get_column_letter(2 + col_num_idx)
                        ws_target.column_dimensions[col_letter_dynamic].width = 20
                    usd_total_column_idx = 2 + num_data_columns
                    ws_target.cell(row=9, column=usd_total_column_idx, value="美元收入").font = Font(color="8B008B")
                    apply_border_to_cell(ws_target.cell(row=9, column=usd_total_column_idx))
                    ws_target.column_dimensions[openpyxl.utils.get_column_letter(usd_total_column_idx)].width = 11
                    sum_for_usd_total_row10 = 0
                    for col_sum_idx in range(num_data_columns):
                        val_row10 = ws_target.cell(row=10, column=2 + col_sum_idx).value
                        if is_number(val_row10):
                            sum_for_usd_total_row10 += float(str(val_row10).replace(',', ''))
                    usd_total_cell_row10 = ws_target.cell(row=10, column=usd_total_column_idx, value=sum_for_usd_total_row10)
                    usd_total_cell_row10.font = Font(color="8B008B")
                    usd_total_cell_row10.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                    apply_border_to_cell(usd_total_cell_row10)
                    for date_row_offset in range(len(max_dates)):
                        current_date_row = 11 + date_row_offset
                        sum_for_usd_date_row = 0
                        for col_sum_idx_date in range(num_data_columns):
                            val_date_row = ws_target.cell(row=current_date_row, column=2 + col_sum_idx_date).value
                            if is_number(val_date_row):
                                sum_for_usd_date_row += float(str(val_date_row).replace(',', ''))
                        usd_date_cell = ws_target.cell(row=current_date_row, column=usd_total_column_idx, value=sum_for_usd_date_row)
                        usd_date_cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                        usd_date_cell.font = Font(color="8B008B")
                        apply_border_to_cell(usd_date_cell)
                    twd_column_idx = usd_total_column_idx + 1
                    ws_target.cell(row=9, column=twd_column_idx, value="台幣收入").font = Font(color="0000FF")
                    apply_border_to_cell(ws_target.cell(row=9, column=twd_column_idx))
                    ws_target.column_dimensions[openpyxl.utils.get_column_letter(twd_column_idx)].width = 11
                    for i in range(1 + len(max_dates)):
                        row_for_twd_calc = 10 + i
                        usd_val_for_twd = ws_target.cell(row=row_for_twd_calc, column=usd_total_column_idx).value
                        twd_calculated_value = 0
                        if is_number(usd_val_for_twd):
                            twd_calculated_value = float(str(usd_val_for_twd).replace(',', '')) * 33
                        twd_cell = ws_target.cell(row=row_for_twd_calc, column=twd_column_idx, value=twd_calculated_value)
                        twd_cell.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                        twd_cell.font = Font(color="0000FF")
                        apply_border_to_cell(twd_cell)
                    for row in range(3, 11):
                        row_title = ws_target[f'A{row}'].value
                        font_color = color_map.get(str(row_title).strip(), None) if row_title else None
                        for col in range(2, 2 + num_data_columns):
                            cell = ws_target.cell(row=row, column=col)
                            if font_color:
                                original_font = cell.font or Font()
                                cell.font = Font(
                                    name=original_font.name,
                                    size=original_font.size,
                                    bold=True,
                                    italic=original_font.italic,
                                    vertAlign=original_font.vertAlign,
                                    underline=original_font.underline,
                                    strike=original_font.strike,
                                    color=font_color
                                )
                    electronic_wallet_sum = 0
                    for col in range(2, 2 + num_data_columns):
                        cell_value = ws_target.cell(row=3, column=col).value
                        if is_number(cell_value):
                            electronic_wallet_sum += float(str(cell_value).replace(',', ''))
                    electronic_wallet_sum_col = usd_total_column_idx
                    ws_target.cell(row=3, column=electronic_wallet_sum_col, value=electronic_wallet_sum)
                    ws_target.cell(row=3, column=electronic_wallet_sum_col).number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                    ws_target.cell(row=3, column=electronic_wallet_sum_col).font = Font(color="00008B", bold=True)
                    apply_border_to_cell(ws_target.cell(row=3, column=electronic_wallet_sum_col))
                    ws_target.cell(row=3, column=electronic_wallet_sum_col + 1, value="←電子錢包總和")
                    ws_target.cell(row=3, column=electronic_wallet_sum_col + 1).font = Font(color="00008B", bold=True)
                    apply_border_to_cell(ws_target.cell(row=3, column=electronic_wallet_sum_col + 1))
                    bonus_storage_sum = 0
                    for col in range(2, 2 + num_data_columns):
                        cell_value = ws_target.cell(row=4, column=col).value
                        if is_number(cell_value):
                            bonus_storage_sum += float(str(cell_value).replace(',', ''))
                    bonus_storage_sum_col = usd_total_column_idx
                    ws_target.cell(row=4, column=bonus_storage_sum_col, value=bonus_storage_sum)
                    ws_target.cell(row=4, column=bonus_storage_sum_col).number_format = FORMAT_NUMBER_COMMA_SEPARATED1
                    ws_target.cell(row=4, column=bonus_storage_sum_col).font = Font(color="8B4513", bold=True)
                    apply_border_to_cell(ws_target.cell(row=4, column=bonus_storage_sum_col))
                    ws_target.cell(row=4, column=bonus_storage_sum_col + 1, value="←獎金暫存總和")
                    ws_target.cell(row=4, column=bonus_storage_sum_col + 1).font = Font(color="8B4513", bold=True)
                    apply_border_to_cell(ws_target.cell(row=4, column=bonus_storage_sum_col + 1))
                    for col in range(2, 2 + num_data_columns):
                        cell = ws_target.cell(row=10, column=col)
                        original_font = cell.font or Font()
                        cell.font = Font(
                            name=original_font.name,
                            size=original_font.size,
                            bold=True,
                            italic=original_font.italic,
                            vertAlign=original_font.vertAlign,
                            underline=original_font.underline,
                            strike=original_font.strike,
                            color="8B008B"
                        )
                        apply_border_to_cell(cell)
                    max_row = 10 + len(max_dates)
                    max_col = twd_column_idx
                    for row in range(2, max_row + 1):
                        for col in range(1, max_col + 1):
                            cell = ws_target.cell(row=row, column=col)
                            apply_border_to_cell(cell)
                    for col_idx in [usd_total_column_idx, twd_column_idx]:
                        max_length = 0
                        for row in range(2, max_row + 1):
                            cell = ws_target.cell(row=row, column=col_idx)
                            value = cell.value
                            if value is None:
                                continue
                            value_str = str(value)
                            if len(value_str) > max_length:
                                max_length = len(value_str)
                        ws_target.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(12, max_length + 2)
                    for col in range(2, 2 + num_data_columns):
                        ws_target.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
                    # 補A9（左右人數）和A10（總計）標題
                    ws_target['A9'] = "左右人數"
                    ws_target['A9'].font = Font(color="006400")
                    apply_border_to_cell(ws_target['A9'])
                    ws_target['A10'] = "總計"
                    ws_target['A10'].font = Font(color="8B008B")
                    apply_border_to_cell(ws_target['A10'])
                wb_target.save(bonus2_path)
                # === 分割報表 ===
                try:
                    import openpyxl
                    from datetime import datetime
                    # 載入剛剛產生的 Bonus2.xlsx
                    workbook = openpyxl.load_workbook(bonus2_path)
                    date_str = datetime.now().strftime("%Y%m%d")
                    for sheet_name in workbook.sheetnames:
                        # 為每個分頁建立新的工作簿
                        new_workbook = openpyxl.Workbook()
                        # 移除預設分頁
                        default_sheet = new_workbook.active
                        new_workbook.remove(default_sheet)
                        # 複製分頁到新的工作簿
                        source_sheet = workbook[sheet_name]
                        new_sheet = new_workbook.create_sheet(title=sheet_name)
                        # 複製欄寬
                        for col in source_sheet.column_dimensions:
                            new_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
                        # 複製列高
                        for row in source_sheet.row_dimensions:
                            new_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
                        # 複製所有儲存格（包含值和格式）
                        for row in source_sheet.rows:
                            for cell in row:
                                new_cell = new_sheet[cell.coordinate]
                                new_cell.value = cell.value
                                self.copy_cell_format(cell, new_cell)
                        # 建立輸出檔名，前面加上日期，並放在 bonus.xlsx 同資料夾
                        split_output_file = os.path.join(bonus_dir, f"{date_str}{sheet_name}.xlsx")
                        new_workbook.save(split_output_file)
                    self.status_label.config(text="✅ 報表生成與分割完成！")
                    # 分割完成後自動打開所在資料夾
                    try:
                        os.startfile(bonus_dir)
                    except Exception:
                        pass
                except Exception as e:
                    self.status_label.config(text=f"報表生成完成，但分割時發生錯誤: {str(e)}")
            except Exception as e:
                self.status_label.config(text=f"錯誤: {str(e)}")
        thread = threading.Thread(target=report_thread)
        thread.daemon = True
        thread.start()

    def count_sheets(self):
        try:
            self.status_label.config(text="正在驗算工作表數量...")
            self.root.update()

            # 獲取所有包含 bonus.xlsx 的資料夾
            bonus_files = glob.glob("output/**/bonus.xlsx", recursive=True)
            if not bonus_files:
                raise Exception("找不到 bonus.xlsx 文件")
            
            # 根據文件修改時間找到最新的 bonus.xlsx
            latest_bonus = max(bonus_files, key=os.path.getmtime)
            
            # 載入工作簿
            wb = openpyxl.load_workbook(latest_bonus)
            
            # 獲取所有工作表名稱
            all_sheets = wb.sheetnames
            
            # 計算實際工作表數量（排除預設的 'Sheet'）
            actual_sheets = [sheet for sheet in all_sheets if sheet.lower() != 'sheet']
            sheet_count = len(actual_sheets)
            
            # 顯示簡潔的結果和路徑
            self.status_label.config(text=f"共:{sheet_count}個帳號工作表(排除sheet)\n{latest_bonus}")
            
        except Exception as e:
            self.status_label.config(text=f"錯誤: {str(e)}")

    def update_latest_path(self):
        """更新最新路徑顯示"""
        try:
            # 獲取所有包含 bonus.xlsx 的資料夾
            bonus_files = glob.glob("output/**/bonus.xlsx", recursive=True)
            if not bonus_files:
                self.path_label.config(text="找不到 bonus.xlsx 文件")
                return
            
            # 根據文件修改時間找到最新的 bonus.xlsx
            latest_bonus = max(bonus_files, key=os.path.getmtime)
            # 只顯示到資料夾層級
            bonus_dir = os.path.dirname(latest_bonus)
            self.path_label.config(text=f"目前路徑：\n{bonus_dir}")
            
        except Exception as e:
            self.path_label.config(text=f"錯誤: {str(e)}")

    def toggle_schedule(self):
        """切換定時狀態"""
        if not self.is_scheduled:
            # 開始定時
            try:
                hour = self.hour_var.get().zfill(2)
                minute = self.minute_var.get().zfill(2)
                
                # 驗證時間格式
                if not (0 <= int(hour) <= 23 and 0 <= int(minute) <= 59):
                    raise ValueError("時間格式不正確")
                
                # 保存時間設定
                self.save_schedule_time(hour, minute)
                
                # 計算等待時間
                now = datetime.now()
                target_time = now.replace(hour=int(hour), minute=int(minute), second=0, microsecond=0)
                if target_time <= now:
                    target_time = target_time.replace(day=target_time.day + 1)
                
                wait_seconds = (target_time - now).total_seconds()
                
                # 設置定時任務
                self.scheduled_task = self.root.after(int(wait_seconds * 1000), self.execute_scheduled_task)
                
                # 更新按鈕狀態
                self.is_scheduled = True
                self.schedule_btn.config(text="取消定時", bg="red")
                self.status_label.config(text=f"已設定自動化定時：[{hour}]:[{minute}]，將在每天此時執行")
                
            except Exception as e:
                self.status_label.config(text=f"設定自動化定時時發生錯誤: {str(e)}")
        else:
            # 取消定時
            if self.scheduled_task:
                self.root.after_cancel(self.scheduled_task)
                self.scheduled_task = None
            
            # 更新按鈕狀態
            self.is_scheduled = False
            self.schedule_btn.config(text="定時啟動", bg="SystemButtonFace")
            self.status_label.config(text="已取消定時")

    def load_schedule_time(self):
        """從文件讀取保存的時間設定"""
        try:
            if os.path.exists("schedule_time.txt"):
                with open("schedule_time.txt", "r") as f:
                    time_str = f.read().strip()
                    hour, minute = time_str.split(":")
                    return hour, minute
        except:
            pass
        return "00", "00"

    def save_schedule_time(self, hour, minute):
        """保存時間設定到文件"""
        try:
            with open("schedule_time.txt", "w") as f:
                f.write(f"{hour}:{minute}")
        except Exception as e:
            self.status_label.config(text=f"保存時間設定時發生錯誤: {str(e)}")

    def execute_scheduled_task(self):
        self.status_label.config(text="開始執行定時任務...")
        def run_automation():
            try:
                self.keyins_running = True
                self.status_label.config(text="開始執行自動化流程...")
                # 執行 keyins
                if os.path.exists("keyins.exe"):
                    os.system("keyins.exe")
                else:
                    os.system("python keyins.py")
                self.status_label.config(text="keyins 執行完成，開始生成報表...")
                self.generate_report()
                # 檢查帳號數量與分頁數量
                try:
                    if not os.path.exists("account.txt"):
                        raise Exception("找不到 account.txt 文件")
                    with open("account.txt", 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    account_count = len(lines) // 3
                    bonus_files = glob.glob("output/**/bonus.xlsx", recursive=True)
                    if not bonus_files:
                        raise Exception("找不到 bonus.xlsx 文件")
                    latest_bonus = max(bonus_files, key=os.path.getmtime)
                    wb = openpyxl.load_workbook(latest_bonus)
                    sheet_count = len([sheet for sheet in wb.sheetnames if sheet.lower() != 'sheet'])
                    if account_count != sheet_count:
                        self.status_label.config(text="帳號數量不符合，請檢查帳號與報表！")
                except Exception as e:
                    self.status_label.config(text=f"檢查帳號數量時發生錯誤: {str(e)}")
            except Exception as e:
                self.status_label.config(text=f"執行自動化流程時發生錯誤: {str(e)}")
            finally:
                self.keyins_running = False
        thread = threading.Thread(target=run_automation)
        thread.daemon = True
        thread.start()

    def start_countdown(self, seconds, callback):
        """開始倒數計時"""
        self.countdown_seconds = seconds
        self.countdown_task = self.root.after(1000, lambda: self.update_countdown(callback))

    def update_countdown(self, callback):
        """更新倒數計時"""
        self.countdown_seconds -= 1
        if self.countdown_seconds > 0:
            # 根據回調函數名稱顯示不同的動作
            action = "執行 keyins"
            if callback.__name__ == "generate_report":
                action = "生成報表"
            elif callback.__name__ == "retry_failed_accounts":
                action = "重試失敗帳號"
            
            self.status_label.config(text=f"倒數 {self.countdown_seconds} 秒後{action}")
            self.countdown_task = self.root.after(1000, lambda: self.update_countdown(callback))
        else:
            callback()

    def check_and_execute_next(self):
        """檢查並執行下一步操作"""
        try:
            # 獲取 account.txt 中的帳號數量
            if not os.path.exists("account.txt"):
                raise Exception("找不到 account.txt 文件")
            
            with open("account.txt", 'r', encoding='utf-8') as f:
                lines = f.readlines()
            account_count = len(lines) // 3

            # 獲取 bonus.xlsx 中的工作表數量
            bonus_files = glob.glob("output/**/bonus.xlsx", recursive=True)
            if not bonus_files:
                raise Exception("找不到 bonus.xlsx 文件")
            
            latest_bonus = max(bonus_files, key=os.path.getmtime)
            wb = openpyxl.load_workbook(latest_bonus)
            sheet_count = len([sheet for sheet in wb.sheetnames if sheet.lower() != 'sheet'])

            # 判斷帳號數量是否符合
            if account_count == sheet_count:
                self.status_label.config(text="帳號數量符合，開始生成報表...")
                self.generate_report()
            else:
                self.status_label.config(text="帳號數量不符合，請檢查帳號與報表！")
                # 單純警告，不做任何自動重試或後續動作

        except Exception as e:
            self.status_label.config(text=f"檢查時發生錯誤: {str(e)}")

    def get_google_drive_file_id(self, url):
        if 'drive.google.com' not in url:
            return url
        if '/file/d/' in url:
            file_id = url.split('/file/d/')[1].split('/')[0]
        elif 'id=' in url:
            file_id = url.split('id=')[1].split('&')[0]
        else:
            return url
        return file_id

    def get_direct_download_url(self, url):
        # 參考 updater.py，將 Google Drive 連結轉 direct download
        if 'drive.google.com' in url:
            file_id = self.get_google_drive_file_id(url)
            return f"https://drive.google.com/uc?id={file_id}"
        return url

    def download_file(self, url, save_path):
        # 參考 updater.py，支援 Google Drive direct download
        direct_url = self.get_direct_download_url(url)
        if 'drive.google.com' in direct_url:
            output = gdown.download(url=direct_url, output=save_path, quiet=False, fuzzy=True, use_cookies=True)
            if output is None:
                raise Exception("gdown 下載失敗")
            return True
        else:
            r = requests.get(direct_url, stream=True)
            r.raise_for_status()
            with open(save_path, 'wb') as f:
                for chunk in r.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
            return True

    def download_version_json(self, url, save_path="temp_version.json"):
        # 參考 updater.py，支援 Google Drive
        if 'drive.google.com' in url:
            file_id = self.get_google_drive_file_id(url)
            gdrive_url = f"https://drive.google.com/uc?id={file_id}"
            output = gdown.download(url=gdrive_url, output=save_path, quiet=True, fuzzy=True, use_cookies=True)
            if output is None:
                raise Exception("gdown 下載失敗")
            with open(save_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            os.remove(save_path)
            return data
        else:
            r = requests.get(url)
            r.raise_for_status()
            return r.json()

    def check_version_on_startup(self):
        try:
            # 下載線上 version.json（支援 Google Drive）
            version_url = "https://drive.google.com/file/d/1oQ20Vx7KOkKH8Bjsbx_LenFBC1_8fXB_/view?usp=drive_link"
            remote_data = self.download_version_json(version_url)
            remote_version = remote_data.get('version', '0.0.0')
            update_msg = remote_data.get('version_message', '')
            download_url = remote_data.get('download_url', '')

            # 讀取本地 version.json
            current_version = '0.0.0'
            try:
                with open('version.json', 'r', encoding='utf-8') as f:
                    current_version = json.load(f).get('version', '0.0.0')
            except Exception:
                pass

            def version_tuple(v):
                return tuple(int(x) for x in v.split('.'))
            def pad_tuple(t, length):
                return t + (0,) * (length - len(t))
            remote_t = version_tuple(remote_version)
            current_t = version_tuple(current_version)
            max_len = max(len(remote_t), len(current_t))
            is_newer = pad_tuple(remote_t, max_len) > pad_tuple(current_t, max_len)

            # debug print
            print("本地版本:", current_version)
            print("遠端版本:", remote_version)
            print("is_newer:", is_newer)
            print("update_btn:", self.update_btn)

            # 取得 direct download 連結
            direct_url = self.get_direct_download_url(download_url)

            if is_newer:
                self.status_label.config(text=f"發現新版本 {remote_version}\n{update_msg}")
                if self.update_btn:
                    self.update_btn.config(state=tk.NORMAL)
                self.new_version_info = remote_data
                self.new_version_info['download_url'] = direct_url  # 後續下載用 direct url
            else:
                self.status_label.config(text="目前已是最新版本")
                if self.update_btn:
                    self.update_btn.config(state=tk.DISABLED)
                self.new_version_info = None
        except Exception as e:
            self.status_label.config(text=f"檢查更新失敗: {str(e)}")
            if self.update_btn:
                self.update_btn.config(state=tk.DISABLED)
            self.new_version_info = None

    def update_version_file(self, version_data):
        with open('version.json', 'w', encoding='utf-8') as f:
            json.dump(version_data, f, indent=4, ensure_ascii=False)

    def download_and_update(self, download_url):
        try:
            temp_zip = "keyins_update.zip"
            self.download_file(download_url, temp_zip)
            with zipfile.ZipFile(temp_zip, 'r') as z:
                z.extractall(".")
            os.remove(temp_zip)
            # 新增：更新本地 version.json
            if self.new_version_info:
                self.update_version_file(self.new_version_info)
            self.status_label.config(text="更新完成，請重新啟動程式！")
        except Exception as e:
            self.status_label.config(text=f"下載或解壓縮失敗: {str(e)}")

    def check_update(self):
        if getattr(self, 'new_version_info', None):
            download_url = self.new_version_info.get("download_url", "")
            if self.update_btn:
                self.update_btn.config(text="下載中...", state=tk.DISABLED)  # 修改狀態文字
            thread = threading.Thread(target=self._threaded_update)
            thread.start()
        else:
            self.status_label.config(text="無可用更新")

    def _threaded_update(self):
        try:
            if not getattr(self, 'new_version_info', None):
                self.status_label.config(text="無可用更新")
                if self.update_btn:
                    self.update_btn.config(text="更新/下載", state=tk.DISABLED)  # 修改狀態文字
                return
            self.download_and_update(self.new_version_info.get("download_url", ""))
            if self.update_btn:
                self.update_btn.config(text="更新/下載", state=tk.DISABLED)  # 修改狀態文字
        except Exception as e:
            self.status_label.config(text=f"更新失敗: {str(e)}")
            if self.update_btn:
                self.update_btn.config(text="更新/下載", state=tk.NORMAL)  # 修改狀態文字

    def run_automation(self):
        if self.keyins_running:
            self.status_label.config(text="keyins 正在執行中，請稍候...")
            return
        def run_automation_thread():
            try:
                self.keyins_running = True
                self.status_label.config(text="開始執行自動化流程...")
                # 執行 keyins
                if os.path.exists("keyins.exe"):
                    os.system("keyins.exe")
                else:
                    os.system("python keyins.py")
                self.status_label.config(text="keyins 執行完成，開始生成報表...")
                self.generate_report()
                # 檢查帳號數量與分頁數量
                try:
                    if not os.path.exists("account.txt"):
                        raise Exception("找不到 account.txt 文件")
                    with open("account.txt", 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    account_count = len(lines) // 3
                    bonus_files = glob.glob("output/**/bonus.xlsx", recursive=True)
                    if not bonus_files:
                        raise Exception("找不到 bonus.xlsx 文件")
                    latest_bonus = max(bonus_files, key=os.path.getmtime)
                    wb = openpyxl.load_workbook(latest_bonus)
                    sheet_count = len([sheet for sheet in wb.sheetnames if sheet.lower() != 'sheet'])
                    if account_count != sheet_count:
                        self.status_label.config(text="帳號數量不符合，請檢查帳號與報表！")
                except Exception as e:
                    self.status_label.config(text=f"檢查帳號數量時發生錯誤: {str(e)}")
            except Exception as e:
                self.status_label.config(text=f"執行自動化流程時發生錯誤: {str(e)}")
            finally:
                self.keyins_running = False
        thread = threading.Thread(target=run_automation_thread)
        thread.daemon = True
        thread.start()

    def open_account(self):
        try:
            if not os.path.exists("account.txt"):
                self.status_label.config(text="錯誤: 找不到 account.txt 文件")
                return
            
            # 使用系統默認程序打開文件
            os.startfile("account.txt")
            self.status_label.config(text="已打開 account.txt")
        except Exception as e:
            self.status_label.config(text=f"錯誤: {str(e)}")

    def open_latest_bonus(self):
        try:
            bonus_files = glob.glob("output/**/bonus.xlsx", recursive=True)
            if not bonus_files:
                raise Exception("找不到 bonus.xlsx 文件")
            latest_bonus = max(bonus_files, key=os.path.getmtime)
            bonus_dir = os.path.dirname(latest_bonus)
            os.startfile(bonus_dir)
            self.status_label.config(text="已打開資料夾")
        except Exception as e:
            self.status_label.config(text=f"錯誤: {str(e)}")

    def open_config(self):
        try:
            if not os.path.exists("config.txt"):
                self.status_label.config(text="錯誤: 找不到 config.txt 文件")
                return
            os.startfile("config.txt")
            self.status_label.config(text="已打開 config.txt")
        except Exception as e:
            self.status_label.config(text=f"錯誤: {str(e)}")

    def open_latest_log(self):
        try:
            log_dirs = glob.glob("logs/*")
            if not log_dirs:
                raise Exception("找不到 logs 目錄")
            latest_log_dir = max(log_dirs, key=os.path.getctime)
            os.startfile(latest_log_dir)
            self.status_label.config(text=f"已打開最新日誌資料夾：{latest_log_dir}")
        except Exception as e:
            self.status_label.config(text=f"錯誤: {str(e)}")

    def copy_cell_format(self, source_cell, target_cell):
        # 複製字體格式
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        # 複製邊框格式
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        # 複製填滿格式
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        # 複製對齊方式
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        # 複製數字格式
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

def ensure_version_json():
    if not os.path.exists('version.json'):
        data = {
            "version": "0.0.0",
            "download_url": "https://drive.google.com/uc?id=1nTMYcWA3tnd-F7FZQ1PfqeGvJjzaJZFs",
            "version_message": "下載keyins"
        }
        with open('version.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

def main():
    ensure_version_json()
    root = tk.Tk()
    app = RetryManagerGUI(root)
    
    # 在 GUI 啟動後，使用背景執行緒檢查更新
    def check_update_in_background():
        def update_thread():
            try:
                # 下載線上 version.json（支援 Google Drive）
                version_url = "https://drive.google.com/file/d/1oQ20Vx7KOkKH8Bjsbx_LenFBC1_8fXB_/view?usp=drive_link"
                remote_data = app.download_version_json(version_url)
                remote_version = remote_data.get('version', '0.0.0')
                update_msg = remote_data.get('version_message', '')
                download_url = remote_data.get('download_url', '')

                # 讀取本地 version.json
                current_version = '0.0.0'
                try:
                    with open('version.json', 'r', encoding='utf-8') as f:
                        current_version = json.load(f).get('version', '0.0.0')
                except Exception:
                    pass

                def version_tuple(v):
                    return tuple(int(x) for x in v.split('.'))
                def pad_tuple(t, length):
                    return t + (0,) * (length - len(t))
                remote_t = version_tuple(remote_version)
                current_t = version_tuple(current_version)
                max_len = max(len(remote_t), len(current_t))
                is_newer = pad_tuple(remote_t, max_len) > pad_tuple(current_t, max_len)

                # 取得 direct download 連結
                direct_url = app.get_direct_download_url(download_url)

                if is_newer:
                    root.after(0, lambda: app.status_label.config(text=f"發現新版本 {remote_version}\n{update_msg}"))
                    root.after(0, lambda: app.update_btn.config(state=tk.NORMAL))
                    app.new_version_info = remote_data
                    app.new_version_info['download_url'] = direct_url  # 後續下載用 direct url
                else:
                    root.after(0, lambda: app.status_label.config(text="目前已是最新版本"))
                    root.after(0, lambda: app.update_btn.config(state=tk.DISABLED))
                    app.new_version_info = None
            except Exception as e:
                root.after(0, lambda: app.status_label.config(text=f"檢查更新失敗: {str(e)}"))
                root.after(0, lambda: app.update_btn.config(state=tk.DISABLED))
                app.new_version_info = None

        # 創建並啟動背景執行緒
        thread = threading.Thread(target=update_thread)
        thread.daemon = True
        thread.start()

    # 使用 after 方法在 GUI 啟動後執行更新檢查
    root.after(100, check_update_in_background)
    
    root.mainloop()

if __name__ == "__main__":
    main()