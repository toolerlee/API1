from flask import Flask, request, jsonify, send_from_directory, Response
import threading
import dropbox
import os
import re
from copy import copy
from openpyxl.styles import Alignment, Font, Border, Side, Color
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
import gc
import csv
from excel_processing_utils import _create_excel_from_csv_files
import requests
import configparser
import json
import ddddocr
from PIL import Image
from io import BytesIO
import time
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict
from datetime import datetime
import glob
from concurrent.futures import ThreadPoolExecutor, as_completed
import random
import pytz

print("=== Flask API 啟動 ===")

app = Flask(__name__)

status = {"running": False, "result": None, "progress": "尚未開始"}
config = {}

tz_taipei = pytz.timezone('Asia/Taipei')

def load_config():
    global config
    parser = configparser.ConfigParser()
    parser.read('config.txt', encoding='utf-8')

    config = {
        'dropbox_app_key': os.getenv('dropbox_app_key') or parser.get('DEFAULT', 'dropbox_app_key', fallback=None),
        'dropbox_app_secret': os.getenv('dropbox_app_secret') or parser.get('DEFAULT', 'dropbox_app_secret', fallback=None),
        'dropbox_refresh_token': os.getenv('dropbox_refresh_token') or parser.get('DEFAULT', 'dropbox_refresh_token', fallback=None),
        'api_action_password': os.getenv('api_action_password') or parser.get('DEFAULT', 'api_action_password', fallback=None),
        'dropbox_account_file_path': os.getenv('dropbox_account_file_path') or parser.get('DEFAULT', 'dropbox_account_file_path', fallback='/Apps/ExcelAPI-app/account/account.txt'),
        'max_concurrent_accounts': parser.get('DEFAULT', 'max_concurrent_accounts', fallback='5'),
        'start_date': parser.get('DEFAULT', 'start_date', fallback='2025/01/01'),
        'end_date': parser.get('DEFAULT', 'end_date', fallback='2025/12/31'),
        'thread_start_delay': parser.get('DEFAULT', 'thread_start_delay', fallback='0.5'),
        'max_login_attempts': parser.get('DEFAULT', 'max_login_attempts', fallback='3'),
        'request_delay': parser.get('DEFAULT', 'request_delay', fallback='2.0'),
        'max_request_retries': parser.get('DEFAULT', 'max_request_retries', fallback='3'),
        'retry_delay': parser.get('DEFAULT', 'retry_delay', fallback='3.0'),
        'dropbox_folder': os.getenv('dropbox_folder') or parser.get('DEFAULT', 'dropbox_folder', fallback='/output'),
        'dropbox_token': None
    }
    
    print('DEBUG: dropbox_app_key =', repr(config['dropbox_app_key']))
    print('DEBUG: dropbox_app_secret =', repr(config['dropbox_app_secret']))
    print('DEBUG: dropbox_refresh_token =', repr(config['dropbox_refresh_token']))
    
    if config['dropbox_refresh_token']:
        try:
            url = "https://api.dropbox.com/oauth2/token"
            data = {"grant_type": "refresh_token", "refresh_token": config['dropbox_refresh_token']}
            auth = (config['dropbox_app_key'], config['dropbox_app_secret'])
            response = requests.post(url, data=data, auth=auth)
            response.raise_for_status()
            config['dropbox_token'] = response.json().get("access_token")
            print("✅ 成功使用 Refresh Token 換取新的 Access Token。")
        except requests.exceptions.RequestException as e:
            print(f"❌ 使用 Refresh Token 換取 Access Token 失敗: {e}")
            if e.response: print(f"    錯誤回應: {e.response.text}")
    else:
        print("⚠️ 警告: 缺少 Dropbox Refresh Token，將無法與 Dropbox 互動。")

    print('RENDER DEBUG: dropbox_app_key =', repr(config['dropbox_app_key']))
    print('RENDER DEBUG: dropbox_app_secret =', repr(config['dropbox_app_secret']))
    print('RENDER DEBUG: dropbox_refresh_token =', repr(config['dropbox_refresh_token']))
    print('DEBUG: 最終 dropbox_token =', repr(config['dropbox_token']))

# 這裡才呼叫 load_config，確保所有函數都已定義
load_config()

global_color_map_for_reports = {
    "紅利積分": "FF0000", "電子錢包": "00008B", "獎金暫存": "8B4513",
    "註冊分": "FF8C00", "商品券": "2F4F4F", "星級": "708090"
}
global_thin_border_for_reports = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

print('RENDER DEBUG: dropbox_app_key =', repr(os.getenv('dropbox_app_key')))
print('RENDER DEBUG: dropbox_app_secret =', repr(os.getenv('dropbox_app_secret')))
print('RENDER DEBUG: dropbox_refresh_token =', repr(os.getenv('dropbox_refresh_token')))
print('RENDER DEBUG: dropbox_account_file_path =', repr(os.getenv('dropbox_account_file_path')))

def is_number_value(value):
    if value is None: return False
    try:
        float(str(value).replace(',', ''))
        return True
    except (ValueError, TypeError):
        return False

def apply_formatting_to_cell(cell, bold=False, font_color_hex=None, border=None, alignment_horizontal='center', alignment_vertical='center'):
    if border: cell.border = border
    cell.alignment = Alignment(horizontal=alignment_horizontal, vertical=alignment_vertical)
    current_font = cell.font if cell.has_style and cell.font else Font()
    new_font_attributes = {
        'name': current_font.name, 'sz': current_font.sz if current_font.sz else 11,
        'b': bold if bold is not None else current_font.b, 'i': current_font.i,
        'vertAlign': current_font.vertAlign, 'underline': current_font.underline,
        'strike': current_font.strike,
    }
    if font_color_hex:
        new_font_attributes['color'] = Color(rgb=font_color_hex)
    elif current_font.color:
        new_font_attributes['color'] = current_font.color
    cell.font = Font(**new_font_attributes)

def copy_cell_format_for_api(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def sort_sheets_by_gold_level_in_api(sheet_names_list, workbook_source):
    def get_sheet_order_key(sheet_name_str):
        ws = workbook_source[sheet_name_str]
        star_level_val = ws['V2'].value
        return 1 if star_level_val and "金級" in str(star_level_val) else 0
    return sorted(sheet_names_list, key=get_sheet_order_key)

def _internal_generate_bonus2_report(source_bonus_xlsx_path, output_bonus2_xlsx_path, result_log):
    result_log.append(f"內部函數：開始生成 Bonus2.xlsx 從 {source_bonus_xlsx_path}")
    try:
        if not os.path.exists(source_bonus_xlsx_path):
            result_log.append(f"❌ 錯誤: 來源 bonus.xlsx '{source_bonus_xlsx_path}' 不存在。")
            return False
        wb_source = openpyxl.load_workbook(source_bonus_xlsx_path, data_only=True)
        wb_target = openpyxl.Workbook()
        if 'Sheet' in wb_target.sheetnames:
            wb_target.remove(wb_target.active)
        # 先建立人名到帳號sheet的 map
        person_sheets_map = defaultdict(list)
        for sheet_name_from_bonus in wb_source.sheetnames:
            name_raw_part = sheet_name_from_bonus.split("_")[0]
            person_identifier = re.sub(r'\d+$', '', name_raw_part)
            person_sheets_map[person_identifier].append(sheet_name_from_bonus)
        # 欄位標題
        headers = ["帳號名稱", "帳號", "紅利積分", "電子錢包", "獎金暫存", "註冊分", "商品券", "星級", "左區人數", "右區人數", "總計"]
        col_map = {
            "紅利積分": 'Q', "電子錢包": 'R', "獎金暫存": 'S', "註冊分": 'T', "商品券": 'U', "星級": 'V',
            "左區人數": 'W', "右區人數": 'X', "總計": 'M'
        }
        for person_id, sheet_names in person_sheets_map.items():
            ws_target = wb_target.create_sheet(title=person_id[:31])
            for col_idx, h in enumerate(headers, 1):
                ws_target.cell(row=1, column=col_idx, value=h)
            row_idx = 2
            for sheet_name in sheet_names:
                ws_src = wb_source[sheet_name]
                # 帳號名稱、帳號
                acc_name = sheet_name.split('_')[0]
                acc_num = sheet_name.split('_')[1] if '_' in sheet_name else ''
                # 取各欄位資料
                row = [acc_name, acc_num]
                for key in ["紅利積分", "電子錢包", "獎金暫存", "註冊分", "商品券", "星級"]:
                    row.append(ws_src[f'{col_map[key]}2'].value)
                row.append(ws_src['W2'].value)  # 左區人數
                row.append(ws_src['X2'].value)  # 右區人數
                row.append(ws_src['M2'].value)  # 總計
                for col_idx, val in enumerate(row, 1):
                    ws_target.cell(row=row_idx, column=col_idx, value=val)
                row_idx += 1
        wb_target.save(output_bonus2_xlsx_path)
        result_log.append(f"✅ Bonus2.xlsx 已成功生成並儲存於 {output_bonus2_xlsx_path}")
        if 'wb_source' in locals(): del wb_source; gc.collect()
        if 'wb_target' in locals(): del wb_target; gc.collect()
        return True
    except Exception as e_gen_b2:
        result_log.append(f"❌ 生成 Bonus2.xlsx 時發生錯誤: {str(e_gen_b2)}")
        print(f"PYTHON_ERROR in _internal_generate_bonus2_report: {e_gen_b2}")
        import traceback
        traceback.print_exc()
        return False

def _internal_split_bonus2_sheets(bonus2_xlsx_path, output_directory_for_split_files, result_log):
    result_log.append(f"內部函數：開始分割 Bonus2.xlsx 從 {bonus2_xlsx_path} 到目錄 {output_directory_for_split_files}")
    split_files_generated_paths = []
    try:
        if not os.path.exists(bonus2_xlsx_path):
            result_log.append(f"❌ 錯誤: Bonus2.xlsx '{bonus2_xlsx_path}' 不存在，無法分割。")
            return []
        workbook_to_split = openpyxl.load_workbook(bonus2_xlsx_path)
        date_str_prefix = datetime.now(tz_taipei).strftime("%Y%m%d")
        if not os.path.exists(output_directory_for_split_files):
            os.makedirs(output_directory_for_split_files, exist_ok=True)
        for sheet_name_to_split in workbook_to_split.sheetnames:
            new_wb_for_sheet = openpyxl.Workbook()
            if new_wb_for_sheet.sheetnames[0] == 'Sheet':
                new_wb_for_sheet.remove(new_wb_for_sheet.active)
            source_sheet_obj = workbook_to_split[sheet_name_to_split]
            target_sheet_in_new_wb = new_wb_for_sheet.create_sheet(title=sheet_name_to_split)
            for col_letter, dim in source_sheet_obj.column_dimensions.items():
                target_sheet_in_new_wb.column_dimensions[col_letter].width = dim.width
            for row_idx, dim in source_sheet_obj.row_dimensions.items():
                target_sheet_in_new_wb.row_dimensions[row_idx].height = dim.height
            for row in source_sheet_obj.iter_rows():
                for cell in row:
                    new_cell = target_sheet_in_new_wb[cell.coordinate]
                    new_cell.value = cell.value
                    if cell.has_style:
                        copy_cell_format_for_api(cell, new_cell)
            split_filename = f"{date_str_prefix}_{sheet_name_to_split}.xlsx"
            full_split_filepath = os.path.join(output_directory_for_split_files, split_filename)
            new_wb_for_sheet.save(full_split_filepath)
            split_files_generated_paths.append(full_split_filepath)
        result_log.append(f"✅ Bonus2.xlsx 已成功按工作表分割。共生成 {len(split_files_generated_paths)} 個檔案。")
        if 'workbook_to_split' in locals():
            del workbook_to_split
            gc.collect()
        return split_files_generated_paths
    except Exception as e_split_b2:
        result_log.append(f"❌ 分割 Bonus2.xlsx 時發生錯誤: {str(e_split_b2)}")
        print(f"PYTHON_ERROR in _internal_split_bonus2_sheets: {e_split_b2}")
        import traceback
        traceback.print_exc()
        return []

def get_random_ua():
    uas = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',]
    return random.choice(uas)

def make_request(session, url, method='get', headers=None, data=None, retry_count=0):
    global config
    request_delay = float(config.get('request_delay', 2.0))
    max_request_retries = int(config.get('max_request_retries', 3))
    retry_delay = float(config.get('retry_delay', 3.0))

    if headers is None: headers = {}
    headers['User-Agent'] = get_random_ua()
    time.sleep(request_delay)
    try:
        if method.lower() == 'get':
            resp = session.get(url, headers=headers, timeout=20)
        else:
            resp = session.post(url, headers=headers, data=data, timeout=20)
        if resp.status_code == 200:
            return resp
        if retry_count < max_request_retries:
            time.sleep(retry_delay)
            return make_request(session, url, method, headers, data, retry_count + 1)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e_req_make:
        if retry_count < max_request_retries:
            time.sleep(retry_delay)
            return make_request(session, url, method, headers, data, retry_count + 1)
        raise Exception(f"請求 {url} 最終失敗 ({type(e_req_make).__name__}): {e_req_make}") from e_req_make

def fetch_account_data_and_save_to_csv(name, user_account_id, user_password, ocr, current_output_dir, headers_for_file):
    global config
    max_login_attempts = int(config.get('max_login_attempts', 3))
    retry_delay = float(config.get('retry_delay', 3.0))
    start_date = config.get('start_date', '2025/01/01')
    end_date = config.get('end_date', '2025/12/31')

    log_folder = os.path.join('logs', datetime.now().strftime('%Y%m%d_%H%M'))
    os.makedirs(log_folder, exist_ok=True)
    retry_log_path = os.path.join(log_folder, 'retry.txt')
    fail_log_path = os.path.join(log_folder, 'fail_log.txt')

    try:
        thread_id = threading.get_ident()
        def log_detail(message):
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
            print(f"[{timestamp}] [Thread-{thread_id}] [Acc: {name}] {message}")
        log_detail("處理開始")

        login_successful = False
        actual_login_attempts = 0
        session = requests.Session()
        for attempt in range(1, max_login_attempts + 1):
            actual_login_attempts = attempt
            log_detail(f"登入嘗試 {attempt}/{max_login_attempts} - 開始")
            
            current_login_url = "https://member.star-rich.net/login"
            current_headers = {"Referer": current_login_url}
            resp_page = make_request(session, current_login_url, headers=current_headers)
            soup_login = BeautifulSoup(resp_page.text, "html.parser")
            
            img_tag = soup_login.find("img", {"id": "MemberLogin1_Image1"})
            if not img_tag or not img_tag.get("src"):
                log_detail(f"登入嘗試 {attempt}: 找不到驗證碼圖片。")
                time.sleep(retry_delay)
                continue
            
            img_url = "https://member.star-rich.net/" + img_tag["src"]
            img_resp = make_request(session, img_url, headers=current_headers)
            img_bytes = img_resp.content
            login_code = ocr.classification(img_bytes)

            if not login_code:
                log_detail(f"登入嘗試 {attempt}: OCR 未返回驗證碼。")
                time.sleep(retry_delay)
                continue
            if login_code[-1] == '4':
                log_detail(f"登入嘗試 {attempt}: 驗證碼 {login_code} 以4結尾。")
                time.sleep(retry_delay)
                continue
            
            login_data = {inp.get("name"): inp.get("value", "") for inp in soup_login.find_all("input") if inp.get("name")}
            login_data.update({
                "MemberLogin1$txtAccound": user_account_id,
                "MemberLogin1$txtPassword": user_password,
                "MemberLogin1$txtCode": login_code,
                "__EVENTTARGET": "MemberLogin1$lkbSignIn", "__EVENTARGUMENT": ""
            })
            login_resp = make_request(session, current_login_url, method='post', headers=current_headers, data=login_data)
            if "登出" in login_resp.text or "歡迎" in login_resp.text:
                login_successful = True
                break
            log_detail(f"登入嘗試 {attempt} 失敗。")
            time.sleep(retry_delay)

        if not login_successful:
            log_detail(f"帳號 {name} ({user_account_id}) 連續 {actual_login_attempts} 次登入失敗。")
            with open(retry_log_path, 'a', encoding='utf-8') as retry_f:
                retry_f.write(f"{name}\\n{user_account_id}\\n{user_password}\\n")
            with open(fail_log_path, 'a', encoding='utf-8') as fail_f:
                fail_f.write(f"{name}_{user_account_id} 連續{actual_login_attempts}次登入失敗\\n")
            raise Exception(f"帳號 {name} ({user_account_id}) 登入失敗")

        home_url = "https://member.star-rich.net/default"
        home_resp = make_request(session, home_url)
        home_soup = BeautifulSoup(home_resp.text, "html.parser")
        h4s = home_soup.select(".h4")
        extra_data_home = [h.text.strip() for h in h4s[:5]]
        while len(extra_data_home) < 5: extra_data_home.append("")
        star_level_tag = home_soup.select_one("#ctl00_cphPageInner_Label_Pin")
        extra_data_home.append(star_level_tag.text.strip() if star_level_tag else "")

        member_url = "https://member.star-rich.net/mem_memlist"
        member_resp = make_request(session, member_url)
        member_soup = BeautifulSoup(member_resp.text, "html.parser")
        left_count_tag = member_soup.select_one("#ctl00_cphPageInner_cphContent_Label_LeftCount")
        right_count_tag = member_soup.select_one("#ctl00_cphPageInner_cphContent_Label_RightCount")
        extra_data_counts = [left_count_tag.text.strip() if left_count_tag else "", right_count_tag.text.strip() if right_count_tag else ""]
        
        all_extra_data_for_row = extra_data_home + extra_data_counts

        bonus_history_url = "https://member.star-rich.net/bonushistory"
        resp_bonus_init = make_request(session, bonus_history_url)
        soup_bonus_init = BeautifulSoup(resp_bonus_init.text, "html.parser")
        viewstate = soup_bonus_init.find("input", {"name": "__VIEWSTATE"})["value"]
        eventvalidation = soup_bonus_init.find("input", {"name": "__EVENTVALIDATION"})["value"]
        viewstategen = soup_bonus_init.find("input", {"name": "__VIEWSTATEGENERATOR"})["value"]
        form_data_bonus_hist = {
            "__EVENTTARGET": "ctl00$cphPageInner$cphContent$Button_Enter", "__EVENTARGUMENT": "",
            "__VIEWSTATE": viewstate, "__VIEWSTATEGENERATOR": viewstategen, "__EVENTVALIDATION": eventvalidation,
            "ctl00$cphPageInner$cphContent$txtStartDate": start_date,
            "ctl00$cphPageInner$cphContent$txtEndDate": end_date,
        }
        response_bonus_page = make_request(session, bonus_history_url, method='post', data=form_data_bonus_hist)
        current_bonus_soup = BeautifulSoup(response_bonus_page.text, "html.parser")
        account_all_data_rows = []
        first_data_row_processed = False
        while True:
            target_table = None
            for tbl_iter in current_bonus_soup.find_all("table"):
                if any("獎金" in th.get_text(strip=True) for th in tbl_iter.find_all("th")):
                    target_table = tbl_iter
                    break
            if not target_table:
                break
            for tr_idx, tr_iter in enumerate(target_table.find_all("tr")):
                if tr_idx == 0:
                    continue
                cols = [td.get_text(strip=True) for td in tr_iter.find_all("td")]
                if cols and "總計" not in cols[0]:
                    bonus_data_part = cols[:-1]
                    if not first_data_row_processed:
                        full_row = bonus_data_part + all_extra_data_for_row
                        first_data_row_processed = True
                    else:
                        full_row = bonus_data_part + [""] * len(all_extra_data_for_row)
                    account_all_data_rows.append(full_row)
            
            next_btn = current_bonus_soup.find(id="ctl00_cphPageInner$cphContent$hpl_Forward")
            if not next_btn or 'disabled' in next_btn.attrs.get('class', []):
                break
            viewstate_next = current_bonus_soup.find("input", {"name": "__VIEWSTATE"})["value"]
            eventvalidation_next = current_bonus_soup.find("input", {"name": "__EVENTVALIDATION"})["value"]
            viewstategen_next = current_bonus_soup.find("input", {"name": "__VIEWSTATEGENERATOR"})["value"]
            form_data_bonus_next = {
                "__EVENTTARGET": "ctl00$cphPageInner$cphContent$hpl_Forward", "__EVENTARGUMENT": "",
                "__VIEWSTATE": viewstate_next, "__VIEWSTATEGENERATOR": viewstategen_next, "__EVENTVALIDATION": eventvalidation_next,
                "ctl00$cphPageInner$cphContent$txtStartDate": start_date,
                "ctl00$cphPageInner$cphContent$txtEndDate": end_date,
            }
            response_bonus_page = make_request(session, bonus_history_url, method='post', data=form_data_bonus_next)
            current_bonus_soup = BeautifulSoup(response_bonus_page.text, "html.parser")

        if not account_all_data_rows and not first_data_row_processed:
            blank_bonus_part = [""] * (len(headers_for_file) - len(all_extra_data_for_row))
            account_all_data_rows.append(blank_bonus_part + all_extra_data_for_row)
            log_detail("無交易紀錄，但已記錄主頁統計數據。")
        
        csv_file_name = f"{name}_{user_account_id}.csv"
        csv_file_path = os.path.join(current_output_dir, csv_file_name)
        save_successful_csv = False
        try:
            with open(csv_file_path, 'w', newline='', encoding='utf-8-sig') as cf:
                writer = csv.writer(cf)
                writer.writerow(headers_for_file)
                if account_all_data_rows:
                    writer.writerows(account_all_data_rows)
            log_detail(f"CSV 儲存成功: {csv_file_path} ({len(account_all_data_rows)} 行)")
            save_successful_csv = True
        except Exception as e_csv_write:
            log_detail(f"❌ CSV 儲存失敗 {csv_file_path}: {e_csv_write}")

        log_detail(f"帳號 {name} ({user_account_id}) 處理完成。")
        return name, user_account_id, save_successful_csv, len(account_all_data_rows) if account_all_data_rows else 0

    except Exception as e_fetch_outer:
        log_detail(f"[CRITICAL WRAPPER] fetch_account_data_and_save_to_csv 發生未處理錯誤: {e_fetch_outer}")
        raise

def main_job():
    global status, config
    print('DEBUG: main_job 啟動時 dropbox_token =', repr(config.get('dropbox_token')))
    status["running"] = True
    status["result"] = "處理中...請稍候..."
    status["progress"] = "初始化中..."
    print("=== main_job 啟動 ===")
    
    result_log = []
    
    folder_name = datetime.now(tz_taipei).strftime('%Y%m%d_%H%M')
    output_dir = os.path.join('資料夾路徑', folder_name)
    try:
        os.makedirs(output_dir, exist_ok=True)
        result_log.append(f"輸出目錄已準備: {output_dir}")
    except Exception as e_mkdir:
        print(f"CRITICAL: 無法創建輸出目錄 {output_dir}: {e_mkdir}")
        status["result"] = f"錯誤: 無法創建輸出目錄 {output_dir}: {e_mkdir}"
        status["progress"] = "目錄創建失敗"
        status["running"] = False
        return

    headers_for_csv_and_excel = [
        "獎金周期", "獎金周期", "消費對等", "經營分紅", "安置獎金", "推薦獎金",
        "消費分紅", "經營對等", "收件中心", "新增加權", "小計", "其他加項",
        "其他減項", "稅額", "補充費", "總計", "紅利積分", "電子錢包",
        "獎金暫存", "註冊分", "商品券", "星級", "左區人數", "右區人數"
    ]

    try:
        ocr_instance = ddddocr.DdddOcr(show_ad=False)
        print("ddddocr.DdddOcr 實例已在 main_job 中成功初始化。")
    except Exception as e_ocr_init:
        print(f"錯誤：在 main_job 中初始化 ddddocr.DdddOcr 失敗: {e_ocr_init}")
        status["result"] = f"錯誤：OCR組件初始化失敗: {e_ocr_init}"
        status["progress"] = "OCR 初始化失敗"
        status["running"] = False
        return

    current_dropbox_token = config.get('dropbox_token', '')
    dropbox_account_file_path_config = config.get('dropbox_account_file_path', '/Apps/ExcelAPI-app/account/account.txt')
    
    accounts = []
    result_log.append(f"準備從 Dropbox 讀取帳號列表: {dropbox_account_file_path_config}")
    if not current_dropbox_token:
        result_log.append("❌ 錯誤: Dropbox Token 未設定，無法讀取帳號檔案。")
        status["result"] = '\\n'.join(result_log)
        status["progress"] = "錯誤: Dropbox Token 未設定"
        status["running"] = False
        return

    try:
        dbx = dropbox.Dropbox(current_dropbox_token)
        _, res = dbx.files_download(path=dropbox_account_file_path_config)
        account_content = res.content.decode('utf-8')
        
        lines = [line.strip() for line in account_content.splitlines() if line.strip()]
        if not lines:
            result_log.append(f"❌ 警告: 從 Dropbox 路徑 {dropbox_account_file_path_config} 下載的帳號檔案內容為空。")
        else:
            for i in range(0, len(lines), 3):
                if i + 2 < len(lines):
                    name, acc, password = lines[i], lines[i+1], lines[i+2]
                    accounts.append((name, acc, password))
            result_log.append(f"✅ 成功從 Dropbox ({dropbox_account_file_path_config}) 讀取並解析了 {len(accounts)} 個帳號。")
    except dropbox.exceptions.ApiError as dbx_err:
        err_msg = f"❌ Dropbox API 錯誤 (讀取帳號檔): {dbx_err}."
        if isinstance(dbx_err.error, dropbox.files.DownloadError) and dbx_err.error.is_path():
            err_msg += " 錯誤細節: 路徑問題 (例如檔案不存在或路徑錯誤)."
        result_log.append(err_msg)
        status["result"] = '\\n'.join(result_log)
        status["progress"] = "錯誤: Dropbox帳號檔讀取失敗"
        status["running"] = False
        return
    except Exception as e_acc_read:
        result_log.append(f"❌ 從 Dropbox 讀取或解析帳號檔案時發生未知錯誤: {e_acc_read}")
        status["result"] = '\\n'.join(result_log)
        status["progress"] = "錯誤: 讀取帳號檔失敗"
        status["running"] = False
        return

    total_accounts = len(accounts)
    success_count = 0
    failed_accounts_info = []
    csv_generation_success_details = []
    csv_generation_failed_details = []

    if total_accounts > 0:
        max_concurrent_accounts = int(config.get('max_concurrent_accounts', 30))
        thread_start_delay = float(config.get('thread_start_delay', 0.5))
        with ThreadPoolExecutor(max_workers=max_concurrent_accounts) as executor:
            futures = []
            for idx, (name_acc, user_id_acc, pass_acc) in enumerate(accounts, 1):
                futures.append(executor.submit(fetch_account_data_and_save_to_csv, name_acc, user_id_acc, pass_acc, ocr_instance, output_dir, headers_for_csv_and_excel))
            time.sleep(thread_start_delay)
            
            completed_count_threads = 0
            for future in as_completed(futures):
                try:
                    acc_name_res, acc_id_res, csv_saved_res, num_rows_res = future.result()
                    if csv_saved_res:
                        success_count += 1
                        csv_generation_success_details.append({'name': acc_name_res, 'id': acc_id_res, 'rows': num_rows_res})
                        result_log.append(f"帳號 {acc_name_res}_{acc_id_res} CSV 資料儲存成功 ({num_rows_res} 行)。")
                    else:
                        csv_generation_failed_details.append({'name': acc_name_res, 'id': acc_id_res, 'reason': 'CSV儲存標記為失敗'})
                        result_log.append(f"帳號 {acc_name_res}_{acc_id_res} CSV 資料儲存失敗 (由函數回報)。")
                        failed_accounts_info.append(f"{acc_name_res}_{acc_id_res} (CSV儲存失敗)")
                except Exception as e_future:
                    msg_future = str(e_future)
                    failed_accounts_info.append(msg_future)
                    result_log.append(f"[錯誤] 執行緒處理時發生錯誤: {msg_future}")
                finally:
                    completed_count_threads += 1
                    status["progress"] = f"CSV處理中: {completed_count_threads}/{total_accounts} (成功儲存CSV: {success_count})"
    else:
        result_log.append("無帳號可處理 (從 Dropbox 讀取的列表為空)。")
        status["progress"] = "完成 (0/0)"

    excel_file_path_local = None
    if success_count > 0:
        target_bonus_xlsx_path = os.path.join(output_dir, "bonus.xlsx")
        try:
            excel_file_path_local = _create_excel_from_csv_files(output_dir, target_bonus_xlsx_path, headers_for_csv_and_excel, result_log.append)
            if excel_file_path_local:
                result_log.append(f"主要的 bonus.xlsx 已成功從 CSV 生成於: {excel_file_path_local}")
            else:
                result_log.append(f"錯誤或警告: 未能從 CSV 檔案生成主要的 bonus.xlsx。")
        except Exception as e_csv_to_excel:
            result_log.append(f"❌ _create_excel_from_csv_files 生成 bonus.xlsx 錯誤: {e_csv_to_excel}")
    else:
        result_log.append("沒有成功生成的 CSV 檔案，跳過 bonus.xlsx 的創建。")
    
    if excel_file_path_local and os.path.exists(excel_file_path_local):
        bonus2_filename = 'Bonus2.xlsx'
        bonus2_file_path_local = os.path.join(output_dir, bonus2_filename)
        if _internal_generate_bonus2_report(excel_file_path_local, bonus2_file_path_local, result_log):
            result_log.append(f"Bonus2.xlsx 已成功生成於: {bonus2_file_path_local}")
            _internal_split_bonus2_sheets(bonus2_file_path_local, output_dir, result_log)
        else:
            result_log.append(f"錯誤或警告: Bonus2.xlsx 未能成功生成。跳過分割。")
    
    dropbox_status_msg_for_summary = ""
    dropbox_folder_for_output = config.get('dropbox_folder', '/output')
    if current_dropbox_token:
        try:
            dbx_reports = dropbox.Dropbox(current_dropbox_token)
            all_files_to_upload = [f for f in os.listdir(output_dir) if f.endswith('.xlsx')]
            base_dropbox_folder = dropbox_folder_for_output.rstrip('/')
            dropbox_target_run_folder = f"{base_dropbox_folder}/{folder_name}"
            uploaded_count = 0
            for f_report in all_files_to_upload:
                local_path_report = os.path.join(output_dir, f_report)
                try:
                    with open(local_path_report, 'rb') as content_report:
                        dbx_path_report = f"{dropbox_target_run_folder}/{f_report}"
                        dbx_reports.files_upload(content_report.read(), dbx_path_report, mode=dropbox.files.WriteMode.overwrite)
                        result_log.append(f"  ✅ 已上傳報表 {f_report} 到 Dropbox ({dbx_path_report})")
                        uploaded_count += 1
                except Exception as e_dbx_upload:
                    result_log.append(f"  ❌ 上傳報表 {f_report} 到 Dropbox 失敗: {e_dbx_upload}")
            dropbox_status_msg_for_summary = f"Dropbox報表: ✅ {uploaded_count} 個檔案已上傳到 {dropbox_target_run_folder}"
        except Exception as e_dbx_generic:
            dropbox_status_msg_for_summary = f"Dropbox報表: ❌ 連接或操作時發生錯誤 - {str(e_dbx_generic)}"
    else:
        dropbox_status_msg_for_summary = "Dropbox報表: ⚠️ Token未設定，跳過上傳"

    final_summary_lines = [f"帳號處理成功: {success_count}/{total_accounts}"]
    if failed_accounts_info:
        final_summary_lines.append(f"帳號處理失敗: {len(failed_accounts_info)}")
        final_summary_lines.append("失敗詳情:")
        for fail_msg in failed_accounts_info:
            final_summary_lines.append(f"  - {fail_msg}")
    
    final_summary_lines.append(dropbox_status_msg_for_summary)
    
    status["result"] = '\\n'.join(final_summary_lines)
    status["progress"] = f"完成: {success_count}/{total_accounts}"
    status["running"] = False
    print("main_job 執行完畢.")

@app.route('/run_main', methods=['POST'])
def run_main():
    # 不再檢查密碼，任何人都能執行主要任務
    if status["running"]:
        return jsonify({"status": "busy", "message": "先前的任務仍在執行中，請稍後再試。"})
    thread = threading.Thread(target=main_job)
    thread.start()
    return jsonify({"status": "started", "message": "主要腳本已啟動執行。"})

@app.route('/status', methods=['GET'])
def get_status():
    return jsonify(status)

@app.route('/')
def serve_index():
    return send_from_directory('.', 'index.html')

@app.route('/version')
def get_version():
    return jsonify({"version": "1.0.0"})

@app.route('/debug-config')
def debug_config():
    # Mask sensitive values before returning
    debug_cfg = {k: ('********' if any(s in k for s in ['token', 'secret', 'password']) else v) for k, v in config.items()}
    return jsonify(debug_cfg)

@app.route('/api/account_file', methods=['GET', 'POST'])
def manage_account_file():
    print('DEBUG: dropbox_token in manage_account_file =', repr(config.get('dropbox_token')))
    if request.method == 'GET':
        dbx_path = config.get('dropbox_account_file_path')
        dbx = dropbox.Dropbox(config.get('dropbox_token'))
        try:
            _, res = dbx.files_download(path=dbx_path)
            content = res.content.decode('utf-8')
            return jsonify({"status": "success", "content": content})
        except dropbox.exceptions.ApiError as err:
            return jsonify({"status": "error", "message": f"無法讀取檔案: {err}"}), 500

    if request.method == 'POST':
        data = request.get_json()
        content = data.get('content')
        dbx_path = config.get('dropbox_account_file_path')
        dbx = dropbox.Dropbox(config.get('dropbox_token'))
        try:
            dbx.files_upload(content.encode('utf-8'), dbx_path, mode=dropbox.files.WriteMode.overwrite)
            return jsonify({"status": "success", "message": "檔案已成功儲存"})
        except dropbox.exceptions.ApiError as err:
            return jsonify({"status": "error", "message": f"無法儲存檔案: {err}"}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 8080)))